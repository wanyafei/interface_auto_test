from openpyxl import load_workbook
from Library import CommonMethod

class excelOpern():
    def __init__(self,excelfile_path,sheet_name):
        '''
        获取指定的文件指定sheet的内容,若文件中数据为空则控制台打印错误信息
        :param excelfile_path:
        :param sheet_name:
        '''
        self.wb=load_workbook(excelfile_path).get_sheet_by_name(sheet_name)
        self.sheetname=sheet_name
    def getcolumndata(self,num):
        '''
        获取指定列的全部内容
        :param num:
        :return:
        '''
        value=[]
        for cell in list(self.wb.columns)[num]:
            if cell.value != None:
                value.append(cell.value)
        return value
    def getrowcolumndata(self,row,column):
        '''
        获取指定行列的区间内容
        :param row:
        :param column:
        :return:
        '''
        for row in range(1,row):
            for colums in range(1,column):
                return self.wb.cell(row=row, column=colums)
    def read_excel_rows_dict(self):
        '''读取Excel 文件不存在则返回报错信息，循环读取行信息与第一行组成jsonarray返回，如果单元格没有数据则返回NONE'''
        max_rows=self.wb.max_row
        max_column=self.wb.max_column
        read_list=[]
        title_list=[]
        total_list=[]
        if max_rows:

            for i in range(0,max_rows):
                rowjson = {}
                for j in range(0,max_column):

                    if i==0:
                        read_list.append(self.wb.cell(row=i+1, column=j+1).value)
                    else:

                        rowjson[read_list[j]]=self.wb.cell(row=i+1, column=j+1).value
                        rowjson["sheetname"] = self.sheetname
                title_list.append(rowjson)
            for list in title_list:
                if list and list["序号"]!= None :
                    total_list.append(list)
            return total_list

        else:
            return u"文件里没有数据"

class openallSheet():
    def __init__(self,excelfile_path):
        '''
        查询Excel中全部sheet的内容
        :param excelfile_path: excel文件的路径
        '''
        self.wb=load_workbook(excelfile_path)
        self.sheets=self.wb.sheetnames
    def getSheetContent(self):
        '''
        读取每一个sheet的数据 将头信息与每一行组成jsonarray返回，如果单元格没有数据则返回NONE
        :return:jsonarray
        '''
        allsheetlist=[]
        for i in range(len(self.sheets)):
            headerlist=[]
            sheet=self.wb[self.sheets[i]]
            max_rows=sheet.max_row
            max_column=sheet.max_column
            sheetlist=[]
            for j in range(0,max_rows):
                rowjson = {}
                for t in range(0,max_column):
                    if j==0:
                        headerlist.append(sheet.cell(row=j+1,column=t+1).value)
                    else:
                        rowjson[headerlist[t]]=sheet.cell(row=j+1,column=t+1).value
                        rowjson["sheetname"]=self.sheets[i]
                sheetlist.append(rowjson)
            for value in sheetlist :
                if value and value["序号"]!= None:
                    allsheetlist.append(value)
        return allsheetlist



if __name__ == '__main__':

    cur_dir=CommonMethod.getPath()
    # wb=openallSheet(cur_dir+"/api_test/api_data/api.xlsx")
    # print(wb.getSheetContent())
    print(excelOpern(cur_dir+"/api_test/api_data/api.xlsx","全息画像").read_excel_rows_dict())

